from openpyxl import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from os import path
from pathlib import Path

class ExportManager:
    """
        ExportManager: Exports data to an excel worksheet
        Fields:
            fpath: the file path to write the worksheet to
        Methods:
            setFilePath(file_path): setter for file_path
            exportToExcel(file_path, rows): writes rows to an excel worksheet
                at file_path
    """


    def __init__(self, gui):
        self.dir = Path.home().as_posix()
        self.fpath = "output.xlsx"
        self.save_path = path.normpath(path.join(self.dir, self.fpath))
        self.gui = gui

    def setFilePath(self, file_path):
        """
            setter for file path attributes
        """
        self.fpath = str(file_path)
        self.save_path = path.normpath(path.join(self.dir, self.fpath))

    def getSavePath(self):
        return self.save_path

    def setDirectory(self, directory):
        self.dir = str(directory)
        self.save_path = path.normpath(path.join(self.dir, self.fpath))


    def exportToExcel(self, rows):
        """
            Writes data to the excel file in fpath.
            rows: 2-dimensional indexed collection (such as a list of lists,
                or tuple of tuples). Each element in the outer collection will
                be read as a row to be written to the worksheet.
        """


        # create workbook
        workb = Workbook()
        works = workb.active

        first_row = ["Starting Component | PIN", "Ending Component | PIN", "Minimum CSA",
                     "Wires", "Splice(s)"]
        works.append(first_row)
        # write rows
        for row in rows:
            works.append(row)
        # save workbook
        dim_holder = DimensionHolder(worksheet=works)

        for col in range(works.min_column, works.max_column + 1):
            dim_holder[get_column_letter(col)] = ColumnDimension(works, min=col, max=col, width=25)

        works.column_dimensions = dim_holder
        print("saving trace to: ", self.save_path)
        workb.save(self.save_path)
