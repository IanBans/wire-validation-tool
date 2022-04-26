import sys
from pathlib import Path
from os.path import dirname, basename
from PySide2.QtWidgets import QWidget, QStackedWidget, QMainWindow, QGridLayout, QLabel
from PySide2.QtWidgets import QFormLayout, QFileDialog, QComboBox, QPushButton, QListWidget
from PySide2.QtWidgets import QApplication, QFrame, QLineEdit, QCheckBox, QHBoxLayout
from PySide2.QtGui import Qt, QPalette, QColor, QBrush
from openpyxl import load_workbook
from inputparser import InputParser
from export import ExportManager
from graphmanager import GraphManager
from csvconfig import CsvConfig


class App(QMainWindow):
    """
       App: the main UI and entry point for the application
       Fields:
        parser: input parser object to get input from files
        graph: GraphManager instance to store data into.
        stacked_widget: the main window container
       Methods:
        setupUI(): creates the UI structure and elements
        readColumnNames(): helper function to fill the column drop down lists
        goToPage(): sets stacked_widget page to target
        setupFilePage(): sets up file picker page
        setupWireReports(): sets up column picker for each report
        readColumnNames(): helper function for wire report dropdowns
    """
    # sets up UI and create parser and GraphManager instances
    def __init__(self):
        super().__init__()
        self.parser = InputParser(self)
        self.graph = GraphManager(self)
        self.export = ExportManager(self)
        self.stacked_widget = QStackedWidget()
        self.wire_report_paths = []
        self.pdc_paths = []
        self.wire_report_configs = CsvConfig("configs.csv", self)
        self.wire_report_list = QListWidget()
        self.left_widget_layout = QFormLayout()
        self.right_widget_layout = QFormLayout()
        self.working_directory = Path.home().as_posix()
        self.setupUI()

    def setupUI(self):
        """
            sets up UI elements and ties them together
        """
        # the stacked widget is the app container that swaps which widget is shown
        # Each widget is a page
        self.stacked_widget.setMinimumSize(500, 300)
        self.stacked_widget.resize(700, 400)
        self.stacked_widget.setWindowTitle('Paccar Wire Validation Tool')
        # self.pages collects all created pages for navigation
        self.pages = {}

        self.setupFilePage()
        self.stacked_widget.show()

        # Qwidget that contains paths of all wire Reports

    def goToPage(self, target):
        """
            target: string representation of the new page the user wants to display
            Changes the visible widget (the current viewable page) shown in the stacked widget
        """
        if target in self.pages:
            self.stacked_widget.setCurrentWidget(self.pages[target])

    # setup for picking all wire reports and PDC
    def setupFilePage(self):
        """
            sets up the file picker page
        """

        def createReportLabel(path, side):
            """
                path: the file the label represents
                side: determine if the label is created for a wire report or pdc
                adds wire report or pdc path label to the view that shows the path
            """

            def removeReport(label):
                """
                    removes the the specified label from the view and deletes the path
                    anywhere it was saved
                """

                if side == "wire":
                    for item in range(self.wire_report_list.count()):
                        if not self.wire_report_list.item(item):
                            continue
                        if self.wire_report_list.item(item).text() == basename(label.text()):
                            self.wire_report_list.takeItem(item)
                    self.wire_report_paths.remove(label.text())
                    self.left_widget_layout.removeRow(label)
                else:
                    self.pdc_paths.remove(label.text())
                    self.right_widget_layout.removeRow(label)
                if self.wire_report_paths and self.pdc_paths:
                    next_button.setEnabled(True)
                else:
                    next_button.setEnabled(False)

            remove_button = QPushButton("Remove")
            remove_button.setMaximumWidth(100)

            label = QLabel(path)
            label.setMinimumWidth(1)
            label.adjustSize()
            remove_button.clicked.connect(lambda: removeReport(label))

            if side == "wire":
                self.left_widget_layout.addRow(label, remove_button)
            else:
                self.right_widget_layout.addRow(label, remove_button)


        def openCSVFileDialog():
            """
                opens the file picker to select .csv files
                to pick fuse maps. Adds fuse maps to path object
                and GUI
            """

            filenames, _ = QFileDialog.getOpenFileNames(self,
                                                       'Choose file',
                                                       self.working_directory,
                                                       'CSV Files (*.csv)')
            for file in filenames:
                if file not in self.pdc_paths:
                    self.pdc_paths.append(file)
                    createReportLabel(file, "pdc")
            if self.wire_report_paths and self.pdc_paths:
                next_button.setEnabled(True)
            else:
                next_button.setEnabled(False)
            if filenames:
                self.working_directory = dirname(filenames[0])


        def openExcelFileDialog():
            """
                opens the file picker sorted to .xlsx files
                to pick wire reports, adds picked files to
                the wire report lists and GUI
            """
            filenames, _ = QFileDialog.getOpenFileNames(self,
                                                       'Choose file',
                                                       self.working_directory,
                                                       'Excel Files (*.xlsx)')

            for file in filenames:
                if file not in self.wire_report_paths:
                    createReportLabel(file, "wire")
                    self.wire_report_paths.append(file)
                    self.wire_report_list.addItem(basename(file))
                self.working_directory = dirname(file)
            if self.wire_report_paths and self.pdc_paths:
                next_button.setEnabled(True)
            else:
                next_button.setEnabled(False)
            if filenames:
                self.working_directory = dirname(filenames[0])

        def openSaveFileDialog():
            """
                opens file picker to choose save location
                adds save path to GUI
            """
            save_file, _ = QFileDialog.getSaveFileName(self,
                                                        "choose save location",
                                                        Path.home().as_posix(),
                                                        'Excel Files (*.xlsx)')
            if save_file:
                self.export.setSavePath(save_file)
                save.setText(self.export.getSavePath())


        file_picker_widgets = QWidget()
        file_picker_layout = QGridLayout()
        file_picker_layout.setColumnStretch(0, 1)
        file_picker_layout.setRowStretch(0, 1)
        file_picker_layout.setColumnStretch(1, 1)
        next_button = QPushButton('Next')
        pdc_button = QPushButton('Add PDC')
        wire_button = QPushButton('Add Wire Reports')
        save = QPushButton("Choose Where to Save ...")
        save.clicked.connect(openSaveFileDialog)
        self.stacked_widget.addWidget(file_picker_widgets)
        file_picker_widgets.setLayout(file_picker_layout)

        self.left_widget_layout.setFormAlignment(Qt.AlignHCenter)
        self.right_widget_layout.setFormAlignment(Qt.AlignHCenter)
        file_picker_layout.addLayout(self.left_widget_layout, 0, 0)
        file_picker_layout.addLayout(self.right_widget_layout, 0, 1)

        # add buttons for wire reports to left side
        self.left_widget_layout.addRow(wire_button)
        wire_button.clicked.connect(openExcelFileDialog)

        # add buttons for pdc to right side
        self.right_widget_layout.addRow(pdc_button)
        pdc_button.clicked.connect(openCSVFileDialog)

        # register current page in page dict
        self.stacked_widget.addWidget(file_picker_widgets)
        self.pages.update({'file_picker': file_picker_widgets})

        # add naviagtion buttons
        next_button.clicked.connect(self.setupWireReports)
        next_button.clicked.connect(lambda: self.goToPage('wire_reports'))
        next_button.setEnabled(False)
        file_picker_layout.addWidget(next_button, 2, 0, 2, 2)
        file_picker_layout.addWidget(save, 1, 0, 1, 2)
        save.setMaximumWidth(200)
        next_button.setMaximumWidth(200)

    def setupWireReports(self):
        """
            creates page to customize column fields for each report
        """
        def makeDict():
            """
                create a new dictionary from combobox_dict
                where it is populated by strings instead of
                combobox objects
                Also checks if data needs to be written to or from csvConfig
            """

            for key, value in combo_box_dict.items():
                # read combox_dict
                report_list = []
                for box in value:
                    report_list.append(box.currentText())

                wire_report_dict.update({key: report_list})
            print(wire_report_dict)

        def saveNewConfig(button_dict, combo_box_dict):
            """
                button_dict key is the index and the value is a list with these values:
                    line: QlineEdit contains the name of the config to be deleted
                    save: save button that is enabled by this method
                    delete: button that deletes config
                    combo_box: conatins all saved csv configs. contents get updated in this method
                combo_boxe_dict: dictionary of wire report names paired with a list of combo boxes
                    that describe what each column is named
                Update the csv file to include another wire report config
                    enables the delete button,and disables the save button and line edit
            """
            index = self.wire_report_list.currentRow()
            new_config = []
            new_config.append(button_dict[index][0].text())
            for key in combo_box_dict.keys():
                if self.wire_report_list.currentItem().text() in key:
                    for value in combo_box_dict[key]:
                        new_config.append(value.currentText())
                    break
            self.wire_report_configs.add(new_config)
            # deactivate buttons to edit and activate button that deletes
            button_dict[index][0].setEnabled(False)
            button_dict[index][1].setEnabled(False)
            button_dict[index][2].setEnabled(True)
            # check if user is overwriting exisiting config
            if button_dict[index][3].findText(button_dict[index][0].text()) != -1:
                return
            # update each combobox holding configs
            for x in button_dict.keys():
                button_dict[x][3].addItem(button_dict[index][0].text())
                button_dict[x][3].update()


        def deleteNewConfig(button_dict):
            """
                index: which wire report is being edited
                button_dict key is the index and the value is a list with these values:
                    line: QlineEdit contains the name of the config to be deleted
                    save: save button that is enabled by this method
                    delete: button that deletes config
                    combo_box: conatins all saved csv configs. contents get updated in this method
                Deletes the user specified saved config from the csv file, disables the delete button,
                and renables the save button and line edit
            """
            index = self.wire_report_list.currentRow()
            self.wire_report_configs.delete(button_dict[index][0].text())

            # deactivate buttons to edit and activate button that deletes
            button_dict[index][0].setEnabled(True)
            button_dict[index][1].setEnabled(True)
            button_dict[index][2].setEnabled(False)

            # update each combobox holding configs
            for x in button_dict.keys():
                box_to_remove = button_dict[x][3].findText(button_dict[index][0].text())
                # if the item is not in the combobox skip
                if box_to_remove == -1:
                    continue
                button_dict[x][3].removeItem(box_to_remove)
                button_dict[x][3].update()

        def loadCsvConfig():
            """
                When the user wants to load a wire report config,
                this method will update all the combo boxes that
                the user would normally fill in themselves
            """
            layer = self.wire_report_list.currentItem().text()
            index = self.wire_report_list.currentRow()
            # ignore if the user selects "choose option"
            if csv_config_buttons[index][3].currentIndex() == 0:
                return
            # look for the correct wire report
            for report in combo_box_dict.keys():
                if layer in report:
                    # load data from csv
                    fields = self.wire_report_configs.search(
                        csv_config_buttons[index][3].currentText())
                    # for each box, change index to new_index
                    counter = 0
                    for box in combo_box_dict[report]:
                        counter += 1
                        new_index = box.findText(fields[counter])
                        if new_index == - 1:
                            print("Error. Attempting to load csv config that is not compatible"
                                   " with current wire report excel sheet")
                            return
                        box.setCurrentIndex(new_index)
                    return

        def sendReports():
            """
                send reports to input Parser
                & print current graph data
            """
            # send pdc to input parser
            for path in self.pdc_paths:
                if path:
                    self.parser.readPDC(path)

            for path, fields in wire_report_dict.items():
                from_tuple = (fields[0], fields[1])
                to_tuple = (fields[2], fields[3])
                self.parser.readReport(path, from_tuple, to_tuple, fields[4], fields[5])

            for _, pdc in self.parser.getPDCs().items():
                self.graph.addPDC(pdc)
            for report in self.parser.getReports():
                self.graph.addReport(report)
            self.export.exportToExcel(self.graph.traceWires())

        # this list contains all the column fields names
        # necessary for reading the input
        fields_list = ['From Component',
                       'From Pin',
                       'To Component',
                       'To Pin',
                       'Wire CSA',
                       'Wire Name']

        page = QWidget()
        page_layout = QGridLayout()
        fields_selector = QStackedWidget()
        submit = QPushButton('Submit')

        # the Dictionary that contains all the wire column fields.
        # The key is the wire path and the value is a list of QComboBoxes
        # that contain the wire report fields
        combo_box_dict = {}
        # same as the above dict but contains strings instead of QCombobox objects
        wire_report_dict = {}

        self.wire_report_list.itemClicked.connect(
            lambda: fields_selector.setCurrentIndex(self.wire_report_list.currentIndex().row()))
        page.setLayout(page_layout)
        self.stacked_widget.addWidget(page)
        self.pages.update({'wire_reports': page})
        page_layout.addWidget(self.wire_report_list, 0, 0)

        page_layout.addWidget(fields_selector, 0, 1, Qt.AlignVCenter)

        # save and load wire harness configurations

        wire_csv_stacked_widget = QStackedWidget()
        # keeps track of the multiple instances of widgets that save and load csv configs
        # {index: [line edit, save buttom, delete button, combo box] }
        csv_config_buttons = {}

        page_layout.addWidget(wire_csv_stacked_widget, 1, 0, 1, 2, Qt.AlignHCenter)
        # change view of page according to which wire report is selected
        self.wire_report_list.itemClicked.connect(
            lambda: wire_csv_stacked_widget.setCurrentIndex(
                self.wire_report_list.currentIndex().row()))

        # create new widget for each wire report that manages wire csv configs
        # Each widget is stacked on each other in wire_csv_stacked_widget
        for x in range(self.wire_report_list.count()):
            container = QWidget()
            wire_csv_stacked_widget.addWidget(container)
            format_selector = QGridLayout()
            container.setLayout(format_selector)

            # create widgets and edit settings
            format_selector.addWidget(PySide2.QtWidgets.QLabel(
                                                    "Or select saved wire report format"), 0, 2)
            combo_box = QComboBox()
            combo_box.addItem("Choose Option")
            for row in self.wire_report_configs.returnAllNames():
                combo_box.addItem(row)
            combo_box.currentIndexChanged.connect(loadCsvConfig)
            instructions = PySide2.QtWidgets.QLabel("To save the above column labels for later "
                "use, please name this configution and press save")
            instructions.adjustSize()
            line = PySide2.QtWidgets.QLineEdit()
            line.setPlaceholderText("Enter the name of this configuration")
            line.setMaximumWidth(400)
            line.adjustSize()
            save_button = QPushButton("Save")
            delete_button = QPushButton("Delete")
            delete_button.setEnabled(False)
            csv_config_buttons.update({x : [line, save_button, delete_button, combo_box]})

            # add widgets to container and fix formatting
            format_selector.addWidget(combo_box, 1, 2)
            format_selector.addWidget(instructions, 0, 0)
            format_selector.addWidget(line, 1, 0)
            button_layout = QHBoxLayout()
            button_layout.addWidget(save_button)
            button_layout.addWidget(delete_button)
            format_selector.addLayout(button_layout, 2, 0)
            format_selector.setSpacing(20)
            format_selector.setMargin(5)
            draw_line = PySide2.QtWidgets.QFrame()
            draw_line.setFrameShape(PySide2.QtWidgets.QFrame.VLine)
            draw_line.setFrameShadow(PySide2.QtWidgets.QFrame.Raised)
            format_selector.addWidget(draw_line, 0, 1, 3, 1)
            container.adjustSize()
            container.setMaximumHeight(container.height())

        # create event listeners for each save and delete button
        for key in csv_config_buttons.keys():
            csv_config_buttons[key][1].clicked.connect(lambda: saveNewConfig(csv_config_buttons, combo_box_dict))
            csv_config_buttons[key][2].clicked.connect(lambda: deleteNewConfig(csv_config_buttons))

        # create the comboboxes that user uses to specify column names
        for wire_report in self.wire_report_paths:
            fields_layout = QFormLayout()
            fields_container = QWidget()
            combo_box_list = []
            fields_container.setLayout(fields_layout)
            combo_box_dict.update({wire_report: combo_box_list})
            column_names = readColumnNames(wire_report)

            for i, _ in enumerate(fields_list):
                combo_box = QComboBox()
                label = QLabel(fields_list[i])
                combo_box.addItem('Choose Option')

                for name in column_names:
                    combo_box.addItem(name)

                combo_box_list.append(combo_box)
                combo_box.setCurrentIndex(0)
                fields_layout.addRow(combo_box, label)

            fields_selector.addWidget(fields_container)
        fields_selector.setCurrentIndex(0)

        # add navigation buttons to the container
        back = QPushButton("Back")
        back.clicked.connect(lambda: self.goToPage("file_picker"))
        home = QPushButton("Home")
        home.clicked.connect(lambda: self.goToPage("file_picker"))
        home.setMinimumWidth(150)
        home_color = QPalette(QApplication.palette())
        home_color.setBrush(QPalette.Button, QBrush(QColor(165, 214, 255)))
        home.setPalette(home_color)
        page_layout.addWidget(home, 3, 1, Qt.AlignRight)
        page_layout.addWidget(back, 2, 0, 1, 1)
        page_layout.addWidget(submit, 3, 0, 1, 1)

        submit.clicked.connect(makeDict)
        submit.clicked.connect(sendReports)

def reportError(error_code):
    """
    error_code: specifies what type of error recieved
    used by other modules to report errors encountered
    """
    print(error_code)


def readColumnNames(filename):
    """
        filename: full file path of report file
        gets first line of worksheet
        and returns it as a list
        helper function for UI drop down boxes
    """
    names = []
    if filename:
        workb = load_workbook(filename, read_only=True)
        sheet = workb.active
        for first_row in sheet.iter_rows(1, 1, 1, sheet.max_column, True):
            names = list(first_row)
    return names


def cleanPathName(path):
    """
        path: a string conatining the full path to a file
        Returns a new string by striping a path name so that only the file name remains

    """
    i = -1
    while path[i] != '/' and path[i] != '\\':
        i = i - 1
    return path[i + 1:]


if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    window = App()
    sys.exit(app.exec_())
