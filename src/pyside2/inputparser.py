import csv, os
from report import Report
from openpyxl import Workbook, load_workbook

'''
   InputParser: Functions to parse and store data from files
   Fields:
    reports: a list of previously parsed report Objects
    pdcs: a dict of previously parsed PDC map lists, with filename keys
   Methods:
    readPDC(filename)
        reads the pdc file specified at filename and inserts it into pdcs
        filename must be unique to other pdc files
    readReport(filename, from_labels, to_labels, csa, desc)
        reads the report specified at filename with the
        column labels
'''
class InputParser:

    def __init__(self):
        self.reports = []
        self.pdcs = {}

    #readPDC(filename)
    #filename: full file path of pdc file
    #Reads the PDC fuse map CSV and returns a list of dictionaries
    #Each element in the list is a row of the PDC file, in dict format
    #{CONNECTOR: (component,pin), FUSE:fuse rating}
    def readPDC(self, filename):

        if filename:
            contents_list = []
            with open(filename, mode='r') as csv_file:
                pdcDict = csv.DictReader(csv_file, delimiter=',')
                print("Successfully opened pdc fuse map..")
                name = os.path.basename(filename)
                for line in pdcDict:
                    contents = {}
                    contents["CONNECTOR"] = (line["CONNECTOR"], line["PIN"])
                    contents["FUSE"] = line["FUSE RATING"]
                    contents_list.append(contents)
                if name not in self.pdcs.keys():
                    self.pdcs[name] = contents_list
                return contents_list

        else:
            print("invalid filename passed to readPDC...")

    #ReadReport(filename, from_labels, to_labels)
    #filename: full file path of report file
    #(from_conn, from_pin): a tuple of strings - the column names for FROM connector and pin
    #(to_conn, to_pin) - a tuple of strings - the column names for TO connector and pin
    #csa - string column name for wire CSA
    #description - string column name for wire description
    #wrapper function for Report class providing storage
    #reads the report, stores the report object in self.reports list, and returns the data
    def readReport(self, filename, from_labels, to_labels, csa, desc):
        report = Report(filename, from_labels, to_labels, csa, desc)
        self.reports.append(report)
        return report

    #getReports()
    #getter to get the current list of saved report objects
    def getReports(self):
        return self.reports

    #ReadReport(filename, from_labels, to_labels)
    #filename: full file path of report file
    #gets first line of worksheet 'filename'
    #and returns it as a list
    #helper function for UI drop down boxes
    def readColumnNames(self, filename):
        if filename:
            wb = load_workbook(filename, read_only=True)
            sheet = wb.active
            names = []
            for first_row in sheet.iter_rows(1, 1, 1, sheet.max_column, True):
                names = list(first_row)

            return names
