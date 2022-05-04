import os
from openpyxl import load_workbook


class Report:

    """
       Report: modeling a single excel wire report file
       Author: Ian Bansenauer
       Properties:
       filepath: the path of the report file
       filename: the name of the original file
       toLabels: a tuple of strings containing the column labels of TO (component, pin)
       fromLabels: a tuple of strings containing the column labels of FROM (component, pin)
       sheet_list: a list output of the report contents in
        {FROM: (component, pin), TO: (component, pin) CSA:int, DESC:str} dictionary format
        read(): reads the report stored in filepath and adds it to sheet_list
    """

    def __init__(self, filepath, from_labels, to_labels, csa, desc, gui):
        self.filepath = filepath
        self.filename = os.path.basename(filepath)
        self.to_labels = to_labels
        self.from_labels = from_labels
        self.csa = csa
        self.desc = desc
        self.sheet_list = []
        self.gui = gui
        self.read()

    def getContents(self):
        """
            getter for sheet_list
        """
        return self.sheet_list

    def read(self):
        """
            reads the report from filepath and stores output in sheet_list, or empty list if
            reading unsuccessful
        """
        try:
            workb = load_workbook(self.filepath, read_only=True)
            sheet = workb.active
            # creates a list of NoneType to check for empty rows
            empty = [None] * sheet.max_column
            try:
                for row in sheet.iter_rows(2, sheet.max_row, values_only=True):
                    row_dict = {}
                    row_list = list(row)
                    # if row is empty, don't process it
                    if row_list != empty:
                        row_dict["FROM"] = (row[self.from_labels[0]],
                                            row[self.from_labels[1]])
                        row_dict["TO"] = (row[self.to_labels[0]],
                                          row[self.to_labels[1]])
                        row_dict["CSA"] = row[self.csa]
                        row_dict["DESC"] = row[self.desc]
                        self.sheet_list.append(row_dict)
                log = "successfully read report: " + str(self.filename)
                print(log)
                self.gui.reportError(log, "log")
                return self.sheet_list
            except KeyError as error:
                log = "check column number " + str(error) + " is correct"
                print(log)
                self.gui.reportError(log, "error")
                return []
        except FileNotFoundError:
            log = "ERROR: could not find a file at" + str(self.filepath)
            print(log, "error")
            return []
